from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "annotation_template.xlsx"

BASE_URL = (
    "https://github.com/pascalr77/Patient/blob/main/"
    "fine_tuning_transcripts_turn_numbered/{tid}.md"
)

START_T, END_T = 1, 49
TURN_START, TURN_END = 5, 20
N_ITEMS = 25


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "annotations"

    # ---- Header ----
    headers = ["transcript_id", "turn"] + [f"item_{i}" for i in range(1, N_ITEMS + 1)]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue-ish
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ---- Data rows ----
    row = 2
    for t in range(START_T, END_T + 1):
        tid = f"T{t}"
        url = BASE_URL.format(tid=tid)

        for turn in range(TURN_START, TURN_END + 1):
            ws.cell(row=row, column=1, value=tid)
            ws.cell(row=row, column=2, value=turn)

            # hyperlink on transcript_id cell
            c = ws.cell(row=row, column=1)
            c.hyperlink = url
            c.style = "Hyperlink"

            # item_1..item_25 left blank
            row += 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 14  # transcript_id
    ws.column_dimensions["B"].width = 8   # turn
    for col_idx in range(3, 3 + N_ITEMS):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Optional: center the turn column
    for r in range(2, row):
        ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(OUTPUT_FILE)
    print(f"✔ Wrote {OUTPUT_FILE} with {(END_T-START_T+1) * (TURN_END-TURN_START+1)} rows")


if __name__ == "__main__":
    main()
